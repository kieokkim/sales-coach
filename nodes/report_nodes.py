import logging
import os
from pathlib import Path

from jinja2 import Environment, FileSystemLoader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

TEMPLATE_DIR = Path(__file__).parent.parent / "templates"
OUTPUT_DIR = Path(__file__).parent.parent / "output"


def build_html_node(state: dict) -> dict:
    errors = list(state.get("errors", []))
    html_report = ""

    try:
        env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
        template = env.get_template("report.html")

        html_report = template.render(
            report_date=state.get("report_date", ""),
            platform_records=state.get("kpi_summary", {}).get("by_platform", []),
            kpi_total=state.get("kpi_total", {}),
            target_summary=state.get("target_summary", {}),
            anomalies=state.get("anomalies", []),
            llm_commentary=state.get("llm_commentary", ""),
        )
        logger.info("HTML 리포트 생성 완료")
    except Exception as e:
        logger.warning(f"HTML 리포트 생성 실패: {e}")
        errors.append(f"HTML 리포트 생성 실패: {e}")

    return {**state, "html_report": html_report, "errors": errors}


def build_excel_node(state: dict) -> dict:
    errors = list(state.get("errors", []))
    excel_path = ""

    try:
        OUTPUT_DIR.mkdir(exist_ok=True)
        report_date = state.get("report_date", "unknown")
        excel_path = str(OUTPUT_DIR / f"sales_report_{report_date}.xlsx")

        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor="2B5797")
        header_font = Font(bold=True, color="FFFFFF")

        # ── 시트 1: KPI_요약 ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "KPI_요약"

        headers = ["채널", "플랫폼", "ZOR", "ZRE", "순영수증", "총매출(원)", "포인트", "봉사료(원)"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        kpi_records = state.get("kpi_summary", {}).get("by_platform", [])
        for row, rec in enumerate(kpi_records, 2):
            ws1.cell(row=row, column=1, value=rec["channel"])
            ws1.cell(row=row, column=2, value=rec["platform"])
            ws1.cell(row=row, column=3, value=rec["zor"])
            ws1.cell(row=row, column=4, value=rec["zre"])
            c5 = ws1.cell(row=row, column=5, value=rec["net_receipt"])
            c5.number_format = '#,##0"건"'
            c6 = ws1.cell(row=row, column=6, value=rec["total_sales"])
            c6.number_format = '#,##0"원"'
            c7 = ws1.cell(row=row, column=7, value=rec["total_point"])
            c7.number_format = '#,##0"P"'
            c8 = ws1.cell(row=row, column=8, value=rec["total_fee"])
            c8.number_format = '#,##0"원"'

        kpi_total = state.get("kpi_total", {})
        total_row = len(kpi_records) + 2
        total_fill = PatternFill("solid", fgColor="DBEAFE")
        total_font = Font(bold=True)
        total_data = [
            "합계", "",
            "", "",
            kpi_total.get("net_receipt", 0),
            kpi_total.get("total_sales", 0),
            kpi_total.get("total_point", 0),
            kpi_total.get("total_fee", 0),
        ]
        fmt_map = {5: '#,##0"건"', 6: '#,##0"원"', 7: '#,##0"P"', 8: '#,##0"원"'}
        for col, val in enumerate(total_data, 1):
            cell = ws1.cell(row=total_row, column=col, value=val)
            cell.fill = total_fill
            cell.font = total_font
            if col in fmt_map:
                cell.number_format = fmt_map[col]

        for col in ws1.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 2
            ws1.column_dimensions[col[0].column_letter].width = min(max_len, 20)

        # ── 시트 2: 타겟달성현황 ──────────────────────────────────────────
        ws_tgt = wb.create_sheet("타겟달성현황")
        tgt_headers = ["플랫폼", "타겟(원)", "누계매출(원)", "달성률(%)", "잔여일수", "필요일평균(원)"]
        for col, h in enumerate(tgt_headers, 1):
            cell = ws_tgt.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        target_summary = state.get("target_summary", {})
        if target_summary and target_summary.get("by_platform"):
            days_remaining = target_summary.get("days_remaining", 0)
            low_fill = PatternFill("solid", fgColor="FFF3CD")
            high_fill = PatternFill("solid", fgColor="D4EDDA")
            for row, pt in enumerate(target_summary["by_platform"], 2):
                pct = pt.get("achievement_pct", 0)
                row_fill = low_fill if pct < 70 else (high_fill if pct >= 90 else None)
                row_vals = [
                    pt.get("platform", ""),
                    pt.get("target", 0),
                    pt.get("actual", 0),
                    pt.get("achievement_pct", 0),
                    days_remaining,
                    pt.get("daily_required", 0),
                ]
                for col, val in enumerate(row_vals, 1):
                    cell = ws_tgt.cell(row=row, column=col, value=val)
                    if row_fill:
                        cell.fill = row_fill
        else:
            ws_tgt.cell(row=2, column=1, value="타겟 미설정")

        for col in ws_tgt.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 2
            ws_tgt.column_dimensions[col[0].column_letter].width = min(max_len, 22)

        # ── 시트 3: 이상치 ────────────────────────────────────────────────
        ws2 = wb.create_sheet("이상치")
        anomaly_headers = ["유형", "플랫폼", "날짜", "메시지", "심각도"]
        for col, h in enumerate(anomaly_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, a in enumerate(state.get("anomalies", []), 2):
            ws2.cell(row=row, column=1, value=a.get("type", ""))
            ws2.cell(row=row, column=2, value=a.get("platform", ""))
            ws2.cell(row=row, column=3, value=a.get("date", ""))
            ws2.cell(row=row, column=4, value=a.get("message", ""))
            ws2.cell(row=row, column=5, value=a.get("severity", ""))

        for col in ws2.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 2
            ws2.column_dimensions[col[0].column_letter].width = min(max_len, 40)

        # ── 시트 4: AI_코멘터리 ───────────────────────────────────────────
        ws3 = wb.create_sheet("AI_코멘터리")
        ws3.cell(row=1, column=1, value="AI 코멘터리").font = Font(bold=True, size=12)
        commentary = state.get("llm_commentary", "")
        ws3.cell(row=2, column=1, value=commentary)
        ws3.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
        ws3.column_dimensions["A"].width = 80
        ws3.row_dimensions[2].height = max(60, len(commentary) // 2)

        wb.save(excel_path)
        logger.info(f"Excel 리포트 저장: {excel_path}")
    except Exception as e:
        logger.warning(f"Excel 리포트 생성 실패: {e}")
        errors.append(f"Excel 리포트 생성 실패: {e}")
        excel_path = ""

    return {**state, "excel_path": excel_path, "errors": errors}
