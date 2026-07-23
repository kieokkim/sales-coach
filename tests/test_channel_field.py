"""channel 필드 도입 회귀테스트 (Decision 33).

배경: 오프라인/온라인 채널구분이 report_nodes.py/pattern_nodes.py/
insight_node.py 세 곳에 매장이름으로 독립 하드코딩돼 있어 신규 매장이
조용히 온라인으로 오분류됐다. kpi_compute_node가 채널을 태깅하고, 세
소비처는 그 channel 필드를 읽는 방식으로 교체했다 — 매장이름 목록이나
SQL IN절 없이도 신규 매장(예: 4번째 가상 매장)이 정상 분류되는지 확인.
"""
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd

from nodes.kpi_nodes import _compute_kpi_for, kpi_compute_node
from nodes.insight_node import _is_monthly_cumulative_issue
from nodes.report_nodes import build_excel_node


def _sample_df(platform: str, sales: int = 1_000_000) -> pd.DataFrame:
    return pd.DataFrame({
        "판매처명": [platform, platform],
        "오더유형": ["ZOR", "ZRE"],
        "매출": [sales, 0],
        "포인트": [0, 0],
        "봉사료": [0, 0],
    })


def test_compute_kpi_for_tags_channel():
    records = _compute_kpi_for(_sample_df("HCC Store 3"), channel="오프라인")
    assert records[0]["channel"] == "오프라인"


def test_kpi_compute_node_tags_new_store_by_source_not_name():
    """신규 매장(HCC Store 3)이 offline_processed에서 왔으면 매장이름 목록에
    없어도 channel='오프라인'로 태깅돼야 한다 — 하드코딩 목록 의존 없음."""
    state = {
        "offline_processed": _sample_df("HCC Store 3"),
        "online_processed": _sample_df("신규 온라인 채널"),
    }
    result = kpi_compute_node(state)
    by_platform = {r["platform"]: r for r in result["kpi_summary"]["by_platform"]}
    assert by_platform["HCC Store 3"]["channel"] == "오프라인"
    assert by_platform["신규 온라인 채널"]["channel"] == "온라인"


def test_monthly_cumulative_gate_recognizes_dynamic_channel_name():
    """channel_names를 그날 실제 kpi_summary에서 뽑아 전달하면, 하드코딩
    목록에 없던 신규 매장명도 '월 누적' 게이트가 정상 작동해야 한다."""
    channel_names = ["HCC Store 3"]
    assert _is_monthly_cumulative_issue(
        "HCC Store 3 월 누적 달성률이 46%로 낮습니다", channel_names
    )
    assert not _is_monthly_cumulative_issue(
        "HCC Store 3에서 오늘 반품이 급증했습니다", channel_names
    )


def test_monthly_cumulative_gate_empty_channel_names_no_match():
    assert not _is_monthly_cumulative_issue(
        "HCC Store 3 월 누적 달성률이 46%로 낮습니다", []
    )


def test_build_excel_node_writes_channel_for_new_store(tmp_path, monkeypatch):
    """report_nodes.py의 엑셀 채널 컬럼이 rec['channel']을 그대로 쓰는지 —
    신규 매장(HCC Store 3)도 하드코딩 목록 없이 '오프라인'으로 나와야 한다."""
    import openpyxl
    import nodes.report_nodes as report_nodes

    monkeypatch.setattr(report_nodes, "OUTPUT_DIR", tmp_path)
    state = {
        "report_date": "20250601",
        "kpi_summary": {
            "by_platform": [
                {
                    "platform": "HCC Store 3", "channel": "오프라인",
                    "zor": 1, "zre": 0, "net_receipt": 1,
                    "total_sales": 100, "total_point": 0, "total_fee": 0,
                },
            ]
        },
    }
    result = build_excel_node(state)
    assert result["errors"] == []

    wb = openpyxl.load_workbook(result["excel_path"])
    ws1 = wb["KPI_요약"]
    assert ws1.cell(row=2, column=1).value == "오프라인"
    assert ws1.cell(row=2, column=2).value == "HCC Store 3"


def test_trend_direction_30d_splits_by_channel_field(tmp_path, monkeypatch):
    """pattern_nodes._trend_direction_30d의 SQL이 platform IN(...) 대신
    channel 컬럼을 읽어 신규 매장도 오프라인 매출에 정상 합산되는지 확인."""
    import nodes.pattern_nodes as pattern_nodes

    db_path = tmp_path / "test.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE daily_kpi (
            report_date TEXT, channel TEXT, platform TEXT, total_revenue REAL
        )
    """)
    dates = [f"2025-06-{d:02d}" for d in range(1, 31)]
    for d in dates:
        conn.execute(
            "INSERT INTO daily_kpi VALUES (?, ?, ?, ?)",
            (d, "오프라인", "HCC Store 3", 1_000_000),
        )
        conn.execute(
            "INSERT INTO daily_kpi VALUES (?, ?, ?, ?)",
            (d, "온라인", "메이크샵", 500_000),
        )
    conn.commit()
    conn.close()

    def fake_get_db():
        return sqlite3.connect(str(db_path))

    monkeypatch.setattr(pattern_nodes, "get_db", fake_get_db)

    result = pattern_nodes._trend_direction_30d({}, "2025-07-01")
    channel_trends = result.get("channel_trends", {})
    assert channel_trends.get("오프라인", {}).get("direction") == "횡보"
    assert channel_trends.get("온라인", {}).get("direction") == "횡보"
